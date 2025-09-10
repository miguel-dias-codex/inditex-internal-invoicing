from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import re

file_path = input("Please enter the path to the Excel (.xlsx) file: ")

# Remove quotation marks at the start and end if present
file_path = file_path.strip('"\'')  # Removes both " and ' from start/end

# Load workbook
wb = load_workbook(file_path)

# 1. Delete sheets named "Items" and "Prices"
for sheet_name in ["Items", "Prices"]:
    if sheet_name in wb.sheetnames:
        std = wb[sheet_name]
        wb.remove(std)

# 2. Format "Orders" sheet
if "Orders" in wb.sheetnames:
    ws = wb["Orders"]
    # Columns to delete (1-based index): E(5), H(8), I(9), J(10), K(11), L(12), M(13), O(15), P(16), Q(17), S(19)
    for col in sorted([19, 17, 16, 15, 13, 12, 11, 10, 9, 8, 5], reverse=True):
        ws.delete_cols(col)

    # Rename Column C from "Customer" to "Department"
    if ws['C1'].value == "Customer":
        ws['C1'].value = "Department"

    # Search and replace in Column C: keep only allowed department names
    allowed_departments = ["Sistemas E-commerce", "Marketing CRM", "CS", "Proyectos"]
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        cell = row[0]
        if isinstance(cell.value, str):
            for dept in allowed_departments:
                if dept in cell.value:
                    cell.value = dept
                    break
            # If no allowed department found, leave the cell

    # Sort all rows (except header) alphabetically by column C
    data = list(ws.iter_rows(min_row=2, values_only=True))
    data_sorted = sorted(data, key=lambda row: (row[2] if row[2] is not None else ""))

    # Overwrite the sheet with sorted data
    for idx, row in enumerate(data_sorted, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=idx, column=col_idx, value=value)

    # Check column B for times (in hh:mm format) and delete those, keeping only the date (in dd.mm.yyyy format)
    time_pattern = re.compile(r'^\d{2}:\d{2}$')
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        cell = row[0]
        if isinstance(cell.value, str) and time_pattern.match(cell.value.strip()):
            cell.value = None
        elif isinstance(cell.value, str) and '.' in cell.value:
            # If cell contains both date and time, keep only the date part
            parts = cell.value.strip().split()
            if len(parts) > 1 and re.match(r'\d{2}\.\d{2}\.\d{4}', parts[0]):
                cell.value = parts[0]

    # Check column G for any cells containing only numericals.
    first_numerical = None
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        cell = row[0]
        if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.isdigit()):
            first_numerical = cell.value
            break

    # If a numerical value was found, paste it in all text cells in column G (except those containing "Proyectos" and G1)
    if first_numerical is not None:
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            cell = row[0]
            if (
                isinstance(cell.value, str)
                and not cell.value.isdigit()
                and "Proyectos" not in cell.value
            ):
                cell.value = first_numerical  # Paste the numerical value found

# 3: Create new sheets

    # Get all unique department names from column C (excluding header and empty cells)
    departments = set()
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        cell_value = row[0].value
        if cell_value and str(cell_value).strip():
            departments.add(cell_value.strip())

    # Get header row values
    headers = [cell.value for cell in ws[1]]

    # For each department, create a new sheet and copy relevant rows
    for dept in departments:
        new_ws = wb.create_sheet(title=dept[:31])  # Sheet names max 31 chars
        # Write header
        for col_idx, header in enumerate(headers, start=1):
            new_ws.cell(row=1, column=col_idx, value=header)
        # Write rows matching the department
        new_row = 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] == dept:
                for col_idx, value in enumerate(row, start=1):
                    new_ws.cell(row=new_row, column=col_idx, value=value)
                new_row += 1

    # Rename the "Orders" sheet to "Total" if it exists
if "Orders" in wb.sheetnames:
    wb["Orders"].title = "Total"

# After creating department sheets (step 3), apply formatting to all sheets
for ws in wb.worksheets:
    # Make text within Row 1 bold, and format the cells to display the whole text within
    for cell in ws[1]:
        cell.font = Font(name='Arial', size=11, bold=True)
    # Set all fonts to Arial size 11 for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name='Arial', size=11, bold=cell.font.bold)

    # Find the first empty cell in column D (4th column)
    row_idx = 2
    while ws.cell(row=row_idx, column=4).value not in (None, ""):
        row_idx += 1

    # Write "Subtotal" in column D for any sheet that is not "Total", otherwise "Total for [Month]"
    if ws.title == "Total":
        label = "Total for [Month]"
    else:
        label = "Subtotal"
    cell_d = ws.cell(row=row_idx, column=4, value=label)
    cell_d.font = Font(name='Arial', size=11, bold=True)
    cell_d.alignment = Alignment(horizontal='right')

    # Write =SUM formula in column E for all values above (from row 2 to row_idx-1), make it bold
    sum_formula = f"=SUM(E2:E{row_idx-1})"
    cell_e = ws.cell(row=row_idx, column=5, value=sum_formula)
    cell_e.font = Font(name='Arial', size=11, bold=True)

    # Write "EUR" in column F, make it bold
    cell_f = ws.cell(row=row_idx, column=6, value="EUR")
    cell_f.font = Font(name='Arial', size=11, bold=True)
    
    # AutoFit column width for all columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

# 4. Save as a copy with "_edited" suffix
folder, filename = os.path.split(file_path)
name, ext = os.path.splitext(filename)
new_filename = f"Codex Report [Month]{ext}"
new_path = os.path.join(folder, new_filename)
wb.save(new_path)

print(f"Edited file saved as: {new_path}")