from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import messagebox
import re

file_path = input("Please enter the path to the Excel (.xlsx) file: ")
# Remove quotation marks at the start and end if present
file_path = file_path.strip('"\'')  # Removes both " and ' from start/end

wb = load_workbook(file_path)

def run_bershka_code():
    # Only work in the sheet named 'Items'
    if "Items" in wb.sheetnames:
        ws = wb["Items"]

        max_row = ws.max_row

        # Find the first empty column in the sheet
        first_empty_col = ws.max_column + 1

        # Copy contents from columns A and C to the first available empty columns
        for row in range(1, max_row + 1):
            if row == 1:
                ws.cell(row=row, column=first_empty_col, value="Order.Item")  # Set header for new column
                ws.cell(row=row, column=first_empty_col + 1, value="Department")  # Set header for new column
            else:
                ws.cell(row=row, column=first_empty_col, value=ws.cell(row=row, column=1).value)  # Copy A
            ws.cell(row=row, column=first_empty_col + 1, value=ws.cell(row=row, column=3).value)  # Copy C

        # Search and replace in the newly created column (first_empty_col + 1): keep only allowed department names
        allowed_departments = ["Sistemas E-commerce", "Marketing CRM", "CS", "Proyectos"]
        for row in ws.iter_rows(min_row=2, min_col=first_empty_col + 1, max_col=first_empty_col + 1):
            cell = row[0]
            if isinstance(cell.value, str):
                for dept in allowed_departments:
                    if dept in cell.value:
                        cell.value = dept
                        break
            # If no allowed department found, leave the cell

        # Copy contents from Column T (20th column, except header) and append to the newly created column (first_empty_col), separated by a period
        for row in range(2, max_row + 1):
            col_new_a = ws.cell(row=row, column=first_empty_col)
            col_t = ws.cell(row=row, column=20)
            if col_t.value:
                if col_new_a.value:
                    col_new_a.value = f"{col_new_a.value}.{col_t.value}"
                else:
                    col_new_a.value = f"{col_t.value}"

    # Save as a copy with "_edited" suffix
    folder, filename = os.path.split(file_path)
    name, ext = os.path.splitext(filename)
    new_filename = f"{name}_edited{ext}"
    new_path = os.path.join(folder, new_filename)
    wb.save(new_path)
    print(f"Edited file saved as: {new_path}")

def run_oysho_code():
    # Only work in the sheet named 'Items'
    if "Items" in wb.sheetnames:
        ws = wb["Items"]

        max_row = ws.max_row

        # Find the first empty column in the sheet
        first_empty_col = ws.max_column + 1

        # Copy contents from columns A and C to the first available empty columns
        for row in range(1, max_row + 1):
            if row == 1:
                ws.cell(row=row, column=first_empty_col, value="Order.Item")  # Set header for new column
                ws.cell(row=row, column=first_empty_col + 1, value="Department")  # Set header for new column
            else:
                ws.cell(row=row, column=first_empty_col, value=ws.cell(row=row, column=1).value)  # Copy A
            ws.cell(row=row, column=first_empty_col + 1, value=ws.cell(row=row, column=3).value)  # Copy C

        # Search and replace in the newly created column (first_empty_col + 1): keep only allowed department names (case insensitive)
        allowed_departments = ["Web", "Producto", "ATC", "Comm", "Marketing", "Merchan", "Etiquetado", "Training", "API"]
        allowed_departments_lower = [dept.lower() for dept in allowed_departments]

        for row in ws.iter_rows(min_row=2, min_col=first_empty_col + 1, max_col=first_empty_col + 1):
            cell = row[0]
            if isinstance(cell.value, str):
                cell_value_lower = cell.value.lower()
                for i, dept_lower in enumerate(allowed_departments_lower):
                    if dept_lower in cell_value_lower:
                        cell.value = allowed_departments[i]
                        break
                else:
                    cell.value = "API"  # Default to "API" if no allowed department

        # Copy contents from Column T (20th column, except header) and append to the newly created column (first_empty_col), separated by a period
        for row in range(2, max_row + 1):
            col_new_a = ws.cell(row=row, column=first_empty_col)
            col_t = ws.cell(row=row, column=20)
            if col_t.value:
                if col_new_a.value:
                    col_new_a.value = f"{col_new_a.value}.{col_t.value}"
                else:
                    col_new_a.value = f"{col_t.value}"

        # After appending, check first_empty_col + 1 for cells containing "API" and update based on Column D (Project name)
        producto_users = {"alejandrojma", "carlabg", "mariacmcas", "danielalsa", "mariaot", "paulalmu", "noeliacgarci", "eliaggpm", "ericamch", "cristinagrib", "ionafa", "brittk"}
        web_users = {"alejandroqma", "diegoco", "guidogpp", "miguelmrav", "meritxellfe", "davidgpra", "carlotafo", "irenesgarci", "marcosls", "sandrafc", "rogersp", "eriksex", "mariatbma", "sylwiadu", "estibalizsil", "paulallop", "clarama"}
        atc_users = {"philipperp", "lauracoso", "annamise"}
        marketing_users = {"claragra", "esterlb", "jorgebper", "judithss", "lauraffe", "nicolei", "oscarvi"}
        comm_users = {"albertolrom", "anaislkd", "annalmiq", "dakotave", "emanuelace", "lauraprm", "mariabesg", "sandrascu"}
        training_users = {"marcalvs", "delfinaf"}

        for row in range(2, max_row + 1):
            dept_cell = ws.cell(row=row, column=first_empty_col + 1)
            project_cell = ws.cell(row=row, column=4)  # Column D (Project name)
            if dept_cell.value == "API" and project_cell.value:
                # Extract username from square brackets
                match = re.search(r"\[([^\[\]]+)\]", str(project_cell.value))
                if match:
                    username = match.group(1).strip().lower()
                else:
                    username = str(project_cell.value).strip().lower()
                if username in producto_users:
                    dept_cell.value = "PRODUCTO"
                elif username in web_users:
                    dept_cell.value = "WEB"
                elif username in atc_users:
                    dept_cell.value = "ATC"
                elif username in marketing_users:
                    dept_cell.value = "MARKETING"
                elif username in comm_users:
                    dept_cell.value = "COMMs"
                elif username in training_users:
                    dept_cell.value = "TRAINING"
                #If Project name contains "Product Translations", set to "PRODUCTO"
                elif "product translations" in str(project_cell.value).lower():
                    dept_cell.value = "PRODUCTO"

        # Check for any remaining "API" cells in first_empty_col + 1
        invalid_projects = set()
        for row in range(2, max_row + 1):
            dept_cell = ws.cell(row=row, column=first_empty_col + 1)
            col_a_cell = ws.cell(row=row, column=1)
            if dept_cell.value == "API" and col_a_cell.value:
                invalid_projects.add(str(col_a_cell.value))

        if invalid_projects:
            error_message = (
                "The following L10N projects don't have a valid user in the project name:\n"
                + "\n".join(invalid_projects)
                + "\nPlease correct the project names in Plunet and export a new report."
            )
            messagebox.showerror("Invalid L10N Projects", error_message)
            raise Exception(error_message)
        
    # Check column first_empty_col + 1 against Column AB (28th column, Target language)
    for row in range(2, max_row + 1):
        dept_cell = ws.cell(row=row, column=first_empty_col + 1)
        target_lang_cell = ws.cell(row=row, column=28)  # Column AB

        dept_value = str(dept_cell.value).strip().lower() if dept_cell.value else ""
        target_lang = str(target_lang_cell.value).strip() if target_lang_cell.value else ""

        # WEB rules
        if dept_value == "web":
            if target_lang in ["eu-ES", "ca", "gl", "es-ES"]:
                dept_cell.value = "WEB ESPAÑA"
            else:
                dept_cell.value = "WEB"

        # PRODUCTO rules
        elif dept_value == "producto":
            if target_lang in ["eu-ES", "ca", "gl", "es-ES"]:
                dept_cell.value = "PRODUCTO ESPAÑA"
            elif target_lang == "ar":
                dept_cell.value = "PRODUCTO ARABIC"
            elif target_lang == "th":
                dept_cell.value = "PRODUCTO THAI"
            else:
                dept_cell.value = "PRODUCTO OTHER"

        # ATC rules
        if dept_value == "atc":
            if target_lang in ["eu-ES", "ca", "gl", "es-ES"]:
                dept_cell.value = "ATC ESPAÑA"
            else:
                dept_cell.value = "ATC"

        # Remaining department rules
        if dept_value in ["comms", "comm"]:
            dept_cell.value = "COMMs"
        elif dept_value == "marketing":
            dept_cell.value = "MARKETING"
        elif dept_value == "merchan":
            dept_cell.value = "MERCHAN"
        elif dept_value == "etiquetado":
            dept_cell.value = "ETIQUETADO"
        elif dept_value == "training":
            dept_cell.value = "TRAINING"

    # Save as a copy with "_edited" suffix
    folder, filename = os.path.split(file_path)
    name, ext = os.path.splitext(filename)
    new_filename = f"{name}_edited{ext}"
    new_path = os.path.join(folder, new_filename)
    wb.save(new_path)
    print(f"Edited file saved as: {new_path}")

def ask_client():
    root = tk.Tk()
    root.title("Select Client")
    root.geometry("300x120")
    label = tk.Label(root, text="Which client are you invoicing?", font=("Arial", 12))
    label.pack(pady=10)

    def bershka_action():
        root.destroy()
        run_bershka_code()

    def oysho_action():
        root.destroy()
        run_oysho_code()

    btn_bershka = tk.Button(root, text="Bershka", width=12, command=bershka_action)
    btn_oysho = tk.Button(root, text="Oysho", width=12, command=oysho_action)
    btn_bershka.pack(pady=5)
    btn_oysho.pack(pady=5)
    root.mainloop()

ask_client()